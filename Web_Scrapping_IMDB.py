from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank','Movie Name','Year of Release','IMDB Rating'])


try:
    source= requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()
    
    soup = BeautifulSoup(source.text,'html.parser') 
    
    movies = soup.find('tbody',{'class':"lister-list"}).find_all('tr')

    for movie in movies:
        
        name = movie.find('td',{'class':"titleColumn"}).a.text
        
        rank = movie.find('td',{'class':"titleColumn"}).get_text(strip=True).split(".")[0]
        
        year = movie.find('td',{'class':"titleColumn"}).span.text.strip('()')
        
        rating = movie.find('td',{'class':"ratingColumn imdbRating"}).strong.text
        
        print(rank,name,year,rating)

        sheet.append([rank,name,year,rating])
            
except Exception as e:
    print(e)

excel.save('Top 250 IMDB Movie Ratings.xlsx')   
        